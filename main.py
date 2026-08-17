#!/usr/bin/env python3
"""考勤统计桌面工具 - 解析钉钉/飞书导出的考勤Excel"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
from datetime import datetime

class AttendanceTool(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("考勤统计工具 v1.0")
        self.geometry("1200x750")
        self.data = None
        self.workdays = None
        self.summary = None
        self.current_file = None
        self._build_ui()

    def _build_ui(self):
        # === Top bar ===
        top = ttk.Frame(self)
        top.pack(fill='x', padx=10, pady=8)

        self.btn_load = ttk.Button(top, text="📂 选择考勤Excel", command=self._load_file)
        self.btn_load.pack(side='left', padx=5)

        self.btn_export = ttk.Button(top, text="📤 导出汇总Excel", command=self._export, state='disabled')
        self.btn_export.pack(side='left', padx=5)

        self.btn_detail = ttk.Button(top, text="🔍 查看明细", command=self._show_detail, state='disabled')
        self.btn_detail.pack(side='left', padx=5)

        self.lbl_file = ttk.Label(top, text="未加载文件", foreground='gray')
        self.lbl_file.pack(side='left', padx=15)

        self.lbl_status = ttk.Label(top, text="", foreground='green')
        self.lbl_status.pack(side='right', padx=10)

        # === Summary cards ===
        card_frame = ttk.Frame(self)
        card_frame.pack(fill='x', padx=10, pady=(0, 8))

        self.cards = {}
        card_titles = ['总人数', '全勤', '迟到最多', '缺勤最多', '加班最多']
        for t in card_titles:
            f = ttk.LabelFrame(card_frame, text=t, padding=(8, 4))
            f.pack(side='left', fill='both', expand=True, padx=3)
            val = ttk.Label(f, text='-', font=('', 16, 'bold'), anchor='center')
            val.pack(fill='both', expand=True)
            self.cards[t] = val

        # === Treeview ===
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=5)

        cols = ['name', 'dept', 'total_days', 'actual_days', 'rate', 'exp_h', 'act_h',
                'in_work_h', 'late_n', 'late_h', 'early_n', 'absent_h',
                'miss_in', 'miss_out', 'makeup', 'absent_d',
                'ot_total', 'ot_pay', 'ot_comp']
        headings = ['姓名', '部门', '应出勤', '实际出勤', '出勤率', '应出勤(h)', '实际(h)',
                    '班内(h)', '迟到次数', '迟到(h)', '早退', '缺勤(h)',
                    '上班缺卡', '下班缺卡', '补卡', '旷工',
                    '加班(h)', '加班费(h)', '调休(h)']

        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings', height=18)
        for c, h in zip(cols, headings):
            self.tree.heading(c, text=h, command=lambda _c=c: self._sort(_c))
            self.tree.column(c, width=75, anchor='center')
        self.tree.column('name', width=100)
        self.tree.column('dept', width=80)

        self.tree.column('total_days', width=65)
        self.tree.column('actual_days', width=65)
        self.tree.column('rate', width=60)
        self.tree.column('exp_h', width=70)
        self.tree.column('act_h', width=70)
        self.tree.column('in_work_h', width=70)
        self.tree.column('late_n', width=65)
        self.tree.column('late_h', width=60)
        self.tree.column('early_n', width=55)
        self.tree.column('absent_h', width=65)
        self.tree.column('miss_in', width=65)
        self.tree.column('miss_out', width=65)
        self.tree.column('makeup', width=55)
        self.tree.column('absent_d', width=55)
        self.tree.column('ot_total', width=70)
        self.tree.column('ot_pay', width=70)
        self.tree.column('ot_comp', width=65)

        scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        self.tree.tag_configure('red', background='#F4CCCC')
        self.tree.tag_configure('orange', background='#FFF2CC')
        self.tree.tag_configure('green', background='#D9EAD3')

        # === Bottom bar ===
        bottom = ttk.Frame(self)
        bottom.pack(fill='x', padx=10, pady=5)
        self.lbl_info = ttk.Label(bottom, text="", foreground='gray', font=('', 9))
        self.lbl_info.pack(side='left')

    def _load_file(self):
        path = filedialog.askopenfilename(
            title="选择考勤Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if not path:
            return
        self._parse(path)

    def _parse(self, path):
        try:
            df_raw = pd.read_excel(path, sheet_name=0, header=None)
        except Exception as e:
            messagebox.showerror("读取失败", f"无法读取文件:\n{e}")
            return

        if len(df_raw) < 3:
            messagebox.showerror("格式错误", "文件数据不足，请检查是否为正确的考勤导出文件")
            return

        data = df_raw.iloc[2:].copy()
        data.columns = ['姓名','工号','部门','组织架构','性别','职务','人员类型','员工状态','合同公司','工作地点','入职日期','离职日期',
            '日期','星期','班次','考勤组',
            '上班打卡时间','上班打卡结果','上班修改原因','上班打卡地点',
            '下班打卡时间','下班打卡结果','下班修改原因','下班打卡地点',
            '应出勤天数','应出勤时长(小时)','休息或未排班天数','实际出勤天数','实际出勤时长(小时)','班内工作时长(小时)','出差天数','外出时长','补卡次数',
            '迟到次数','迟到时长(小时)','严重迟到次数','严重迟到时长(小时)','早退次数','早退时长(小时)','缺勤时长(小时)','上班缺卡次数','下班缺卡次数','旷工天数',
            '加班总时长(小时)','加班总时长计加班费(小时)','加班总时长计调休(小时)']

        num_cols = ['应出勤天数','应出勤时长(小时)','休息或未排班天数','实际出勤天数','实际出勤时长(小时)','班内工作时长(小时)',
            '出差天数','外出时长','补卡次数','迟到次数','迟到时长(小时)','严重迟到次数','严重迟到时长(小时)',
            '早退次数','早退时长(小时)','缺勤时长(小时)','上班缺卡次数','下班缺卡次数','旷工天数',
            '加班总时长(小时)','加班总时长计加班费(小时)','加班总时长计调休(小时)']
        for c in num_cols:
            data[c] = pd.to_numeric(data[c], errors='coerce').fillna(0)

        self.data = data
        self.workdays = data[data['应出勤天数'] > 0]
        self.current_file = path

        # Build summary
        names = sorted([n for n in data['姓名'].dropna().unique() if pd.notna(n) and n != '姓名'])
        rows = []
        for name in names:
            ndf = self.workdays[self.workdays['姓名'] == name]
            if len(ndf) == 0:
                continue
            dept = ndf['部门'].iloc[0]
            total_wd = int(ndf['应出勤天数'].sum())
            actual_d = int(ndf['实际出勤天数'].sum())
            rate = f"{actual_d/total_wd*100:.0f}%" if total_wd > 0 else "0%"
            rows.append({
                'name': name, 'dept': dept,
                'total_days': total_wd, 'actual_days': actual_d, 'rate': rate,
                'exp_h': round(ndf['应出勤时长(小时)'].sum(), 1),
                'act_h': round(ndf['实际出勤时长(小时)'].sum(), 1),
                'in_work_h': round(ndf['班内工作时长(小时)'].sum(), 1),
                'late_n': int(ndf['迟到次数'].sum()),
                'late_h': round(ndf['迟到时长(小时)'].sum(), 1),
                'early_n': int(ndf['早退次数'].sum()),
                'absent_h': round(ndf['缺勤时长(小时)'].sum(), 1),
                'miss_in': int(ndf['上班缺卡次数'].sum()),
                'miss_out': int(ndf['下班缺卡次数'].sum()),
                'makeup': int(ndf['补卡次数'].sum()),
                'absent_d': int(ndf['旷工天数'].sum()),
                'ot_total': round(ndf['加班总时长(小时)'].sum(), 1),
                'ot_pay': round(ndf['加班总时长计加班费(小时)'].sum(), 1),
                'ot_comp': round(ndf['加班总时长计调休(小时)'].sum(), 1),
            })

        self.summary = pd.DataFrame(rows)
        self._populate_table()
        self._update_cards()

        fname = os.path.basename(path)
        self.lbl_file.config(text=f"📄 {fname}")
        self.lbl_status.config(text="✅ 加载成功")
        self.btn_export.config(state='normal')
        self.btn_detail.config(state='normal')

        # Parse date range
        dates = pd.to_datetime(self.workdays['日期'], errors='coerce').dropna()
        if len(dates) > 0:
            period = f"{dates.min().strftime('%Y/%m/%d')} ~ {dates.max().strftime('%Y/%m/%d')}"
        else:
            period = "未知"
        self.lbl_info.config(text=f"统计周期: {period}  |  总人数: {len(names)}  |  工作日记录: {len(self.workdays)} 条")
        self.after(2000, lambda: self.lbl_status.config(text=""))

    def _populate_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for _, r in self.summary.iterrows():
            tag = 'green'
            if r['absent_d'] > 0 or r['rate'] != '100%':
                try:
                    rate_val = int(r['rate'].replace('%',''))
                    if rate_val < 80 or r['absent_d'] > 0:
                        tag = 'red'
                    else:
                        tag = 'orange'
                except:
                    tag = 'green'

            vals = [r[c] for c in ['name','dept','total_days','actual_days','rate',
                'exp_h','act_h','in_work_h','late_n','late_h','early_n','absent_h',
                'miss_in','miss_out','makeup','absent_d','ot_total','ot_pay','ot_comp']]
            self.tree.insert('', 'end', values=vals, tags=(tag,))

    def _update_cards(self):
        total = len(self.summary)
        full_att = len(self.summary[self.summary['rate'] == '100%'])

        most_late = self.summary.loc[self.summary['late_n'].idxmax()]
        most_absent = self.summary.loc[self.summary['absent_d'].idxmax()]
        most_ot = self.summary.loc[self.summary['ot_total'].idxmax()]

        self.cards['总人数'].config(text=str(total))
        self.cards['全勤'].config(text=f"{full_att}人")
        self.cards['迟到最多'].config(text=f"{most_late['name']}\n{int(most_late['late_n'])}次")
        self.cards['缺勤最多'].config(text=f"{most_absent['name']}\n{int(most_absent['absent_d'])}天")
        self.cards['加班最多'].config(text=f"{most_ot['name']}\n{most_ot['ot_total']}h")

    def _sort(self, col):
        items = [(self.tree.set(i, col), i) for i in self.tree.get_children('')]
        try:
            items.sort(key=lambda x: float(x[0].replace('%','').replace('人','').replace('次','').replace('天','').replace('h','')))
        except:
            items.sort(key=lambda x: x[0])

        # Toggle order
        if hasattr(self, '_sort_col') and self._sort_col == col and self._sort_asc:
            items.reverse()
            self._sort_asc = False
        else:
            self._sort_asc = True
        self._sort_col = col

        for idx, (_, iid) in enumerate(items):
            self.tree.move(iid, '', idx)

    def _show_detail(self):
        if self.data is None:
            return
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("提示", "请先在表格中选中要查看的人员")
            return

        name = self.tree.item(sel[0])['values'][0]
        ndf = self.data[self.data['姓名'] == name].sort_values('日期', ascending=False)

        win = tk.Toplevel(self)
        win.title(f"📋 {name} - 打卡明细")
        win.geometry("900x500")

        cols = ['日期','星期','班次','上班打卡','上班结果','下班打卡','下班结果',
                '出勤(h)','班内(h)','迟到','早退','缺勤(h)','补卡','加班(h)']
        tree = ttk.Treeview(win, columns=cols, show='headings', height=20)
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=90, anchor='center')
        tree.column('日期', width=100)
        tree.column('上班打卡', width=100)
        tree.column('下班打卡', width=100)

        scrollbar = ttk.Scrollbar(win, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')

        for _, r in ndf.iterrows():
            vals = [
                str(r['日期']), r['星期'], r['班次'],
                str(r['上班打卡时间']), str(r['上班打卡结果']),
                str(r['下班打卡时间']), str(r['下班打卡结果']),
                round(r['实际出勤时长(小时)'],1), round(r['班内工作时长(小时)'],1),
                int(r['迟到次数']), int(r['早退次数']),
                round(r['缺勤时长(小时)'],1), int(r['补卡次数']),
                round(r['加班总时长(小时)'],1)
            ]
            tag = 'green'
            if r['迟到次数'] > 0 or r['上班缺卡次数'] > 0 or r['下班缺卡次数'] > 0:
                tag = 'orange'
            if r['旷工天数'] > 0:
                tag = 'red'
            tree.insert('', 'end', values=vals, tags=(tag,))

        tree.tag_configure('red', background='#F4CCCC')
        tree.tag_configure('orange', background='#FFF2CC')
        tree.tag_configure('green', background='#D9EAD3')

    def _export(self):
        if self.summary is None:
            return
        path = filedialog.asksaveasfilename(
            title="保存汇总Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="考勤统计汇总.xlsx"
        )
        if not path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "考勤汇总"

            headers = ['姓名','部门','应出勤天数','实际出勤天数','出勤率',
                '应出勤时长(h)','实际时长(h)','班内时长(h)',
                '迟到次数','迟到时长(h)','早退次数','缺勤时长(h)',
                '上班缺卡','下班缺卡','补卡次数','旷工天数',
                '加班总时长(h)','加班费(h)','调休(h)']

            title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
            title_fill = PatternFill('solid', fgColor='2F5496')
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill('solid', fgColor='4472C4')
            data_font = Font(name='微软雅黑', size=10)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
            center = Alignment(horizontal='center', vertical='center')

            ws.merge_cells('A1:S1')
            ws['A1'] = f'考勤统计汇总 - {datetime.now().strftime("%Y年%m月%d日")}'
            ws['A1'].font = title_font
            ws['A1'].fill = title_fill
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 35

            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border

            red_fill = PatternFill('solid', fgColor='F4CCCC')
            orange_fill = PatternFill('solid', fgColor='FFF2CC')
            green_fill = PatternFill('solid', fgColor='D9EAD3')

            for i, (_, r) in enumerate(self.summary.iterrows(), 3):
                vals = [r['name'], r['dept'], r['total_days'], r['actual_days'], r['rate'],
                    r['exp_h'], r['act_h'], r['in_work_h'], r['late_n'], r['late_h'],
                    r['early_n'], r['absent_h'], r['miss_in'], r['miss_out'],
                    r['makeup'], r['absent_d'], r['ot_total'], r['ot_pay'], r['ot_comp']]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=i, column=c, value=v)
                    cell.font = data_font
                    cell.alignment = center
                    cell.border = thin_border

                try:
                    rate_val = int(str(r['rate']).replace('%',''))
                    if r['absent_d'] > 0 or rate_val < 80:
                        fill = red_fill
                    elif rate_val < 100:
                        fill = orange_fill
                    else:
                        fill = green_fill
                    for c in range(1, len(headers)+1):
                        ws.cell(row=i, column=c).fill = fill
                except:
                    pass

            wb.save(path)
            self.lbl_status.config(text="✅ 导出成功")
            self.after(2000, lambda: self.lbl_status.config(text=""))
        except Exception as e:
            messagebox.showerror("导出失败", str(e))


if __name__ == '__main__':
    app = AttendanceTool()
    app.mainloop()
